import tkinter as tk
import tkinter.ttk as ttk
import sqlite3
import time
import os

from tkinter import messagebox, simpledialog
from ttkthemes import ThemedTk
from tkinter import filedialog

from win32 import win32print
import win32ui
import openpyxl
from openpyxl import Workbook

# pip install pywin32
# pip install openpyxl

edit_mode = True


def main():
    root = ThemedTk(theme="vista")
    root.title('College Student Management')

    style = ttk.Style()
    style.theme_use("vista")

    root.tk_setPalette(background="#f3f3f3")

    def authenticate_user():
        while True:
            username = simpledialog.askstring(
                "Authentication", "Enter username:")
            if username is None:
                root.destroy()
                return
            password = simpledialog.askstring(
                "Authentication", "Enter password:", show="*")
            if password is None:
                root.destroy()
                return

            if username == 'admin' and password == 'admin':
                return
            else:
                messagebox.showerror(
                    "Authentication Failed", "Invalid username or password. Please try again.")

    authenticate_user()

    column_indices = {
        'ID': 1,
        'First Name': 2,
        'Last Name': 3,
        'Full Name': 4,
        'Enrollment Number': 5,
        'Contact Number': 6
    }

    button_frame = ttk.Frame(root)
    button_frame.pack(pady=(10, 10))

    def create_student():
        first_name = first_name_entry.get()
        last_name = last_name_entry.get()
        enrollment_number = enrollment_number_entry.get()
        contact_number = contact_number_entry.get()

        if not (first_name and last_name and enrollment_number and contact_number):
            messagebox.showerror("Error", "All fields must be filled out.")
            return

        full_name = f"{first_name} {last_name}"
        cursor.execute('INSERT INTO students (first_name, last_name, full_name, enrollment_number, contact_number) VALUES (?, ?, ?, ?, ?)',
                       (first_name, last_name, full_name, enrollment_number, contact_number))

        conn.commit()

        clear_entry_widgets()

        display_students()

        create_database_backup(cursor, conn,  'student.db', 'backup_directory')

    def display_students():
        for record in treeview.get_children():
            treeview.delete(record)

        cursor.execute('SELECT * FROM students')
        records = cursor.fetchall()

        for i, record in enumerate(records):
            tags = ('oddrow', 'evenrow')[i % 2]
            treeview.insert('', 'end', values=record, tags=tags)

    def edit_student():
        selected_record = treeview.selection()

        if not selected_record:
            return

        student_data = treeview.item(selected_record)['values']

        edited_data = simpledialog.askstring("Edit Student", "Edit Student Data (First Name, Last Name, Full Name, Enrollment Number, Contact Number)",
                                             initialvalue=f"{student_data[1]}, {student_data[2]}, {student_data[3]}, {student_data[4]}, {student_data[5]}")

        if edited_data:
            edited_data = tuple(edited_data.split(", "))

            if not all(edited_data):
                messagebox.showerror("Error", "All fields must be filled out.")
                return

            cursor.execute('UPDATE students SET first_name=?, last_name=?, full_name=?, enrollment_number=?, contact_number=? WHERE id=?',
                           (*edited_data, student_data[0]))

            conn.commit()

            display_students()

            create_database_backup(
                cursor, conn,  'student.db', 'backup_directory')

    def delete_student():
        selected_record = treeview.selection()

        if not selected_record:
            return

        cursor.execute('DELETE FROM students WHERE id=?',
                       (treeview.item(selected_record)['values'][0],))

        conn.commit()
        clear_entry_widgets()
        display_students()

        create_database_backup(cursor, conn,  'student.db', 'backup_directory')

    def delete_all_students():
        confirmation = messagebox.askyesno(
            "Confirmation", "Are you sure you want to delete all student records?\nThis action cannot be undone.")

        if confirmation:
            delete_all_button.config(state=tk.DISABLED)
            root.update()
            # time.sleep(5)

            delete_all_button.config(state=tk.NORMAL)

            cursor.execute('DELETE FROM students')
            conn.commit()

            clear_entry_widgets()
            display_students()

            create_database_backup(
                cursor, conn,  'student.db', 'backup_directory')

    def clear_entry_widgets():
        first_name_entry.delete(0, tk.END)
        last_name_entry.delete(0, tk.END)
        enrollment_number_entry.delete(0, tk.END)
        contact_number_entry.delete(0, tk.END)

    def clear_search():
        search_entry.delete(0, tk.END)
        display_students()

    def search_students():
        selected_column = search_column_var.get()
        search_text = search_entry.get().strip()
        search_window = tk.Toplevel(root)
        search_window.title('Search Results')

        search_treeview = ttk.Treeview(search_window, columns=(
            'ID', 'First Name', 'Last Name', 'Full Name', 'Enrollment Number', 'Contact Number'), show="headings", height=20)
        search_treeview.heading('#1', text='ID')
        search_treeview.heading('#2', text='First Name')
        search_treeview.heading('#3', text='Last Name')
        search_treeview.heading('#4', text='Full Name')
        search_treeview.heading('#5', text='Enrollment Number')
        search_treeview.heading('#6', text='Contact Number')
        search_treeview.pack()

        cursor.execute(
            f'SELECT * FROM students WHERE {selected_column} LIKE ?', (f'%{search_text}%',))
        records = cursor.fetchall()

        for record in records:
            search_treeview.insert('', 'end', values=record)

        print_button = ttk.Button(
            search_window, text='Print', command=lambda: print_results(search_treeview))
        print_button.pack()

    def print_results(search_treeview):
        printer_name = win32print.GetDefaultPrinter()
        printer_info = win32print.GetPrinter(printer_name, 2)

        hprinter = win32print.OpenPrinter(printer_name)
        printer_dc = win32ui.CreateDC()
        printer_dc.CreatePrinterDC(printer_name)

        printer_dc.StartDoc('Search Results')
        printer_dc.StartPage()

        results = search_treeview.get_children()

        x = 100
        y = 100

        for result in results:
            values = search_treeview.item(result, 'values')
            text = ' - '.join(values)
            printer_dc.TextOut(x, y, text)
            y += 20

        printer_dc.EndPage()
        printer_dc.EndDoc()
        printer_dc.DeleteDC()
        win32print.ClosePrinter(hprinter)

    selected_column = tk.StringVar()

    def clear_column_selection():
        column_listbox.selection_clear(0, tk.END)

    def print_selected_data():
        selected_columns = column_listbox.curselection()
        if selected_columns:
            selected_columns = [column_listbox.get(
                i) for i in selected_columns]

            selected_data_window = tk.Toplevel(root)
            selected_data_window.title(
                f'Selected Data: {", ".join(selected_columns)}')

            selected_treeview = ttk.Treeview(
                selected_data_window, columns=selected_columns, show="headings", height=20)

            for column in selected_columns:
                selected_treeview.heading(column, text=column)

            selected_treeview.pack()

            for record in treeview.get_children():
                values = treeview.item(record, 'values')
                selected_values = [values[column_indices[column] - 1]
                                   for column in selected_columns]
                selected_treeview.insert('', 'end', values=selected_values)

            print_button = ttk.Button(
                selected_data_window, text='Print', command=lambda: print_results(selected_treeview))
            print_button.pack()

    def import_data_from_excel():
        excel_file_path = filedialog.askopenfilename(
            filetypes=[("Excel Files", "*.xlsx")])

        if excel_file_path:
            try:
                workbook = openpyxl.load_workbook(excel_file_path)
                sheet = workbook.active

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if len(row) < 4:
                        messagebox.showerror(
                            "Error", "Invalid data format in the Excel file. Each row should have at least 4 columns of data.")
                        return

                    first_name = row[0]
                    last_name = row[1]
                    enrollment_number = row[2]
                    contact_number = row[3]

                    full_name = f"{first_name} {last_name}"

                    cursor.execute('INSERT INTO students (first_name, last_name, full_name, enrollment_number, contact_number) VALUES (?, ?, ?, ?, ?)',
                                   (first_name, last_name, full_name, enrollment_number, contact_number))

                conn.commit()

                messagebox.showinfo("Success", "Data imported successfully!")

                display_students()

                create_database_backup(
                    cursor, conn, 'student.db', 'backup_directory')

            except Exception as e:
                messagebox.showerror(
                    "Error", f"An error occurred during import:\n{str(e)}")

    def export_data_to_excel():
        workbook = Workbook()
        sheet = workbook.active

        headers = ['ID', 'First Name', 'Last Name',
                   'Full Name', 'Enrollment Number', 'Contact Number']
        sheet.append(headers)

        cursor.execute('SELECT * FROM students')
        records = cursor.fetchall()
        for record in records:
            sheet.append(record)

        excel_file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if excel_file_path:
            workbook.save(excel_file_path)
            messagebox.showinfo("Export Successful",
                                f"Data exported to {excel_file_path} successfully!")

    def toggle_edit_mode():
        global edit_mode
        if edit_mode:
            edit_mode = False
            create_button.config(state=tk.DISABLED)
            edit_button.config(state=tk.DISABLED)
            delete_button.config(state=tk.DISABLED)
            delete_all_button.config(state=tk.DISABLED)
            import_button.config(state=tk.DISABLED)
            export_button.config(state=tk.DISABLED)
            enable_editing_button.config(text='Enable Editing')
        else:
            credentials = simpledialog.askstring(
                "Authentication", "Enter username and password separated by space (e.g., username password):")
            if credentials is None:
                return

            username, password = credentials.split()

            if username == 'editing' and password == 'editing':
                edit_mode = True
                create_button.config(state=tk.NORMAL)
                edit_button.config(state=tk.NORMAL)
                delete_button.config(state=tk.NORMAL)
                delete_all_button.config(state=tk.NORMAL)
                import_button.config(state=tk.NORMAL)
                export_button.config(state=tk.NORMAL)
                enable_editing_button.config(text='Disable Editing')
            else:
                messagebox.showerror(
                    "Authentication Failed", "Invalid username or password. Please try again.")

    conn = sqlite3.connect('student.db')
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY,
            first_name TEXT,
            last_name TEXT,
            full_name TEXT,
            enrollment_number TEXT,
            contact_number TEXT
        )
    ''')

    conn.commit()

    create_backup_table(cursor)
    create_backup_trigger(cursor, 'students')

    backup_directory = 'backup_directory'

    first_name_label = ttk.Label(root, text='First Name:', anchor="w")
    first_name_label.pack()
    first_name_entry = ttk.Entry(root, width=40)
    first_name_entry.pack()

    last_name_label = ttk.Label(root, text='Last Name:', anchor="w")
    last_name_label.pack()
    last_name_entry = ttk.Entry(root, width=40)
    last_name_entry.pack()

    full_name_label = ttk.Label(root, text='Full Name:', anchor="w")
    full_name_label.pack()
    full_name_entry = ttk.Entry(root, width=40)
    full_name_entry.pack()

    full_name_label.destroy()
    full_name_entry.destroy()

    enrollment_number_label = ttk.Label(
        root, text='Enrollment Number:', anchor="w")
    enrollment_number_label.pack()
    enrollment_number_entry = ttk.Entry(root, width=40)
    enrollment_number_entry.pack()

    contact_number_label = ttk.Label(root, text='Contact Number:', anchor="w")
    contact_number_label.pack()
    contact_number_entry = ttk.Entry(root, width=40)
    contact_number_entry.pack()

    button_frame = ttk.Frame(root)
    button_frame.pack(pady=(10, 10))

    create_button = ttk.Button(
        button_frame, text='Create', command=create_student)
    create_button.pack(side=tk.LEFT, padx=10)

    edit_button = ttk.Button(button_frame, text='Edit', command=edit_student)
    edit_button.pack(side=tk.LEFT, padx=10)

    delete_button = ttk.Button(
        button_frame, text='Delete', command=delete_student)
    delete_button.pack(side=tk.LEFT, padx=10)

    delete_all_button = ttk.Button(
        button_frame, text='Delete All', command=delete_all_students)
    delete_all_button.pack(side=tk.LEFT, padx=10)

    import_button = ttk.Button(
        button_frame, text='Import from Excel', command=import_data_from_excel)
    import_button.pack(side=tk.LEFT, padx=10)

    export_button = ttk.Button(
        button_frame, text='Export to Excel', command=export_data_to_excel)
    export_button.pack(side=tk.LEFT, padx=10)

    enable_editing_button = ttk.Button(
        button_frame, text='Enable Editing', command=toggle_edit_mode)
    enable_editing_button.pack(side=tk.LEFT, padx=10)

    new_backup_button = ttk.Button(
        button_frame, text='Create New Backup', command=lambda: create_database_backup(cursor, conn, 'student.db', 'backup_directory'))
    new_backup_button.pack(side=tk.LEFT, padx=10)

    search_frame = ttk.Frame(root)
    search_frame.pack(pady=(10, 10))

    search_label = ttk.Label(search_frame, text='Search Text:')
    search_label.grid(row=0, column=0, padx=5, sticky='e')

    search_entry = ttk.Entry(search_frame, width=30)
    search_entry.grid(row=0, column=1, padx=5)

    search_column_label = ttk.Label(search_frame, text='Search Column:')
    search_column_label.grid(row=0, column=2, padx=5, sticky='e')

    search_columns = ['first_name', 'last_name',
                      'full_name', 'enrollment_number', 'contact_number']

    search_column_var = tk.StringVar()

    search_column_dropdown = ttk.Combobox(
        search_frame, textvariable=search_column_var, values=search_columns)
    search_column_dropdown.grid(row=0, column=3, padx=5)
    search_column_dropdown.set('first_name')

    search_button = ttk.Button(
        search_frame, text='Search', command=search_students)
    search_button.grid(row=0, column=4, padx=5)

    clear_search_button = ttk.Button(
        search_frame, text='Clear Search', command=clear_search)
    clear_search_button.grid(row=0, column=5, padx=5)

    column_selection_frame = ttk.Frame(root)
    column_selection_frame.pack(pady=(10, 10))

    column_listbox = tk.Listbox(column_selection_frame, selectmode=tk.MULTIPLE)
    column_listbox.pack(side=tk.LEFT, padx=5)

    for column_name in column_indices.keys():
        column_listbox.insert(tk.END, column_name)

    clear_selection_button = ttk.Button(
        column_selection_frame, text='Clear Selection', command=clear_column_selection)
    clear_selection_button.pack(side=tk.LEFT, padx=5)

    print_selected_button = ttk.Button(
        column_selection_frame, text='Print Selected', command=print_selected_data)
    print_selected_button.pack(side=tk.LEFT, padx=5)

    tree_frame = ttk.Frame(root)
    tree_frame.pack(side="left", fill="both", expand=True)

    treeview = ttk.Treeview(tree_frame, columns=('ID', 'First Name', 'Last Name',
                            'Full Name', 'Enrollment Number', 'Contact Number'), show='headings', height=20)
    treeview.heading('#1', text='ID')
    treeview.heading('#2', text='First Name')
    treeview.heading('#3', text='Last Name')
    treeview.heading('#4', text='Full Name')
    treeview.heading('#5', text='Enrollment Number')
    treeview.heading('#6', text='Contact Number')

    scrollbar = ttk.Scrollbar(
        tree_frame, orient="vertical", command=treeview.yview)
    treeview.configure(yscrollcommand=scrollbar.set)

    scrollbar.pack(side="right", fill="y")
    treeview.pack(side="left", fill="both", expand=True)

    below_tree_frame = ttk.Frame(root)
    below_tree_frame.pack()

    display_students()
    toggle_edit_mode()
    root.mainloop()


def create_database_backup(cursor, conn, db_file, backup_dir):
    try:
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)

        backup_datetime = time.strftime('%d-%m-%Y_%I-%M-%p')
        backup_file = f"{backup_dir}/{backup_datetime}_backup.db"

        if not os.path.exists(backup_file):
            conn_backup = sqlite3.connect(backup_file)
            cursor_backup = conn_backup.cursor()

            cursor_backup.execute('''
                CREATE TABLE IF NOT EXISTS students (
                    id INTEGER PRIMARY KEY,
                    first_name TEXT,
                    last_name TEXT,
                    full_name TEXT,
                    enrollment_number TEXT,
                    contact_number TEXT
                )
            ''')
            conn_backup.commit()
        else:
            conn_backup = sqlite3.connect(backup_file)
            cursor_backup = conn_backup.cursor()

        cursor.execute('SELECT * FROM students')
        records = cursor.fetchall()

        for record in records:
            cursor_backup.execute(
                'INSERT INTO students VALUES (?, ?, ?, ?, ?, ?)', record)

        conn_backup.commit()
        conn_backup.close()

        cursor.execute(
            f"INSERT INTO backups (table_name, action) VALUES ('database', 'UPDATE')")
        conn.commit()

        print(f"Database data appended to {backup_file}")

        excel_file = f"{backup_dir}/{backup_datetime}_backup.xlsx"
        workbook = Workbook()
        sheet = workbook.active

        headers = ['ID', 'First Name', 'Last Name',
                   'Full Name', 'Enrollment Number', 'Contact Number']
        sheet.append(headers)

        conn_backup = sqlite3.connect(backup_file)
        cursor_backup = conn_backup.cursor()
        cursor_backup.execute('SELECT * FROM students')
        backup_records = cursor_backup.fetchall()

        for backup_record in backup_records:
            sheet.append(backup_record)

        workbook.save(excel_file)
        print(f"Database data saved as Excel file: {excel_file}")

    except Exception as e:
        print(f"Error creating/updating backup: {str(e)}")


def create_backup_table(cursor):
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS backups (
            id INTEGER PRIMARY KEY,
            table_name TEXT,
            action TEXT
        )
    ''')


def create_backup_trigger(cursor, table_name):
    cursor.execute(f'''
        CREATE TRIGGER IF NOT EXISTS {table_name}_backup_trigger
        AFTER INSERT ON {table_name}
        BEGIN
            INSERT INTO backups (table_name, action) VALUES ('{table_name}', 'INSERT');
        END;
    ''')

    cursor.execute(f'''
        CREATE TRIGGER IF NOT EXISTS {table_name}_update_backup_trigger
        AFTER UPDATE ON {table_name}
        BEGIN
            INSERT INTO backups (table_name, action) VALUES ('{table_name}', 'UPDATE');
        END;
    ''')

    cursor.execute(f'''
        CREATE TRIGGER IF NOT EXISTS {table_name}_delete_backup_trigger
        AFTER DELETE ON {table_name}
        BEGIN
            INSERT INTO backups (table_name, action) VALUES ('{table_name}', 'DELETE');
        END;
    ''')


if __name__ == "__main__":
    main()
